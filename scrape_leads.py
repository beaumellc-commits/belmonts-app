"""
Belmonts — Scraper de leads en Île-de-France (Phase 1, local).

Scrape Pages Jaunes + Google Maps avec Playwright et génère un fichier
Excel sur ton Bureau, avec les téléphones, adresses, etc.

USAGE
─────
    python3 scrape_leads.py                      # toute IDF (~60 min)
    python3 scrape_leads.py --zone 75            # Paris uniquement (~25 min)
    python3 scrape_leads.py --zone 75,92,93,94   # petite couronne (~50 min)
    python3 scrape_leads.py --test               # 1 arrondissement (~3 min)
    python3 scrape_leads.py --no-google          # Pages Jaunes seulement
    python3 scrape_leads.py --headless           # navigateur invisible
    python3 scrape_leads.py --output ~/Desktop/mes_leads.xlsx

REPRISE APRÈS INTERRUPTION
──────────────────────────
Le scraper sauve sa progression dans `.cache/leads_scraping.json`. Si tu
fais Ctrl+C, tu relances la même commande, il reprend où il s'était arrêté.
Pour repartir de zéro : supprime le dossier `.cache/`.
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import random
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

# ─── ZONES & QUERIES ──────────────────────────────────────────────────────────
ZONES: dict[str, list[str]] = {
    "75": [f"Paris {i}" for i in range(1, 21)],
    "92": [
        "Neuilly-sur-Seine", "Boulogne-Billancourt", "Nanterre", "Courbevoie",
        "Levallois-Perret", "Colombes", "Asnières-sur-Seine", "Clichy",
        "Issy-les-Moulineaux", "Rueil-Malmaison",
    ],
    "93": [
        "Saint-Denis", "Montreuil", "Aubervilliers", "Noisy-le-Grand",
        "Aulnay-sous-Bois", "Rosny-sous-Bois", "Bobigny", "Pantin",
    ],
    "94": [
        "Créteil", "Vincennes", "Ivry-sur-Seine", "Vitry-sur-Seine",
        "Champigny-sur-Marne", "Saint-Maur-des-Fossés", "Maisons-Alfort",
    ],
    "78": [
        "Versailles", "Saint-Germain-en-Laye", "Poissy", "Mantes-la-Jolie",
        "Sartrouville", "Vélizy-Villacoublay",
    ],
    "77": ["Melun", "Meaux", "Chelles", "Torcy", "Pontault-Combault"],
    "91": ["Évry", "Corbeil-Essonnes", "Palaiseau", "Massy", "Savigny-sur-Orge"],
    "95": ["Cergy", "Argenteuil", "Sarcelles", "Pontoise", "Franconville"],
}

QUERIES = [
    {"label": "Syndic", "term": "syndic de copropriété"},
    {"label": "Agence", "term": "agence immobilière gestion locative"},
]

CACHE_DIR = Path(".cache")
CACHE_FILE = CACHE_DIR / "leads_scraping.json"

PHONE_RE = re.compile(r"(?:\+33\s?|0)[1-9](?:[\s.\-]?\d{2}){4}")
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


# ─── UTILITIES ────────────────────────────────────────────────────────────────
def slugify(text: str) -> str:
    """Normalise un nom pour la déduplication."""
    text = unicodedata.normalize("NFKD", text or "")
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_phone(raw: str) -> str:
    """Reformate un tél FR en '01 23 45 67 89'."""
    if not raw:
        return ""
    digits = re.sub(r"[^\d+]", "", raw)
    if digits.startswith("+33"):
        digits = "0" + digits[3:]
    if not (digits.startswith("0") and len(digits) == 10):
        m = PHONE_RE.search(raw)
        if not m:
            return ""
        digits = re.sub(r"[^\d]", "", m.group(0))
        if len(digits) != 10:
            return ""
    return " ".join([digits[0:2], digits[2:4], digits[4:6], digits[6:8], digits[8:10]])


def load_cache() -> dict[str, Any]:
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text())
        except Exception:
            pass
    return {"done": [], "leads": []}


def save_cache(cache: dict[str, Any]) -> None:
    CACHE_DIR.mkdir(exist_ok=True)
    CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2))


def polite_sleep(a: float = 1.5, b: float = 3.5) -> None:
    time.sleep(random.uniform(a, b))


# ─── PAGES JAUNES ─────────────────────────────────────────────────────────────
def scrape_pages_jaunes(page, term: str, ville: str, dept: str, type_label: str) -> list[dict]:
    """Scrape Pages Jaunes pour (term, ville). Récupère le tél via clic + base64."""
    leads: list[dict] = []
    url = (
        "https://www.pagesjaunes.fr/annuaire/chercherlespros"
        f"?quoiqui={term.replace(' ', '+')}&ou={ville.replace(' ', '+')}"
    )

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        polite_sleep(1.0, 2.0)

        # Accepte les cookies si la bannière apparaît
        for sel in ["#didomi-notice-agree-button", "button:has-text('Accepter')", "#onetrust-accept-btn-handler"]:
            try:
                btn = page.query_selector(sel)
                if btn:
                    btn.click()
                    polite_sleep(0.5, 1.0)
                    break
            except Exception:
                pass

        page.wait_for_selector("article, .bi-bloc, .bi", timeout=10000)
    except Exception as e:
        print(f"   ⚠️  PJ load error {ville}/{term}: {e}")
        return leads

    # Clic sur tous les boutons "Afficher le numéro" pour révéler les tels
    try:
        page.evaluate("""
            () => {
                document.querySelectorAll('.show-button, .pj-show-number, [class*="show-number"], button[class*="bi-tel"]')
                  .forEach(b => b.click());
            }
        """)
        polite_sleep(0.8, 1.5)
    except Exception:
        pass

    # Scroll pour déclencher lazy loading
    for _ in range(3):
        page.mouse.wheel(0, 1500)
        polite_sleep(0.5, 1.0)

    cards = page.query_selector_all("article.bi, li.bi, .bi-bloc")
    if not cards:
        cards = page.query_selector_all("article")

    for card in cards[:25]:
        lead = _parse_pj_card(card, ville, dept, type_label)
        if lead:
            leads.append(lead)
    return leads


def _parse_pj_card(card, ville: str, dept: str, type_label: str) -> dict | None:
    # Nom
    nom_el = (
        card.query_selector("a.bi-denomination")
        or card.query_selector(".denomination-links")
        or card.query_selector("h3.denomination-content")
        or card.query_selector("h2 a")
        or card.query_selector("h3")
    )
    if not nom_el:
        return None
    nom = (nom_el.inner_text() or "").strip()
    if not nom or len(nom) < 2:
        return None

    text_blob = ""
    try:
        text_blob = card.inner_text() or ""
    except Exception:
        pass

    # Téléphone — stratégie 1 : data-pjlb (base64)
    tel = ""
    try:
        pjlb = card.get_attribute("data-pjlb")
        if pjlb:
            data = json.loads(pjlb)
            tel_b64 = data.get("tel") or data.get("telephone") or ""
            if tel_b64:
                # Padding base64 si nécessaire
                pad = "=" * (-len(tel_b64) % 4)
                try:
                    decoded = base64.b64decode(tel_b64 + pad).decode("utf-8", errors="ignore")
                    tel = normalize_phone(decoded)
                except Exception:
                    pass
    except Exception:
        pass

    # Stratégie 2 : sélecteurs visibles (après clic sur "Afficher")
    if not tel:
        for sel in [
            "a[href^='tel:']",
            ".coord-numero",
            ".num-tel",
            "span.bi-phone",
            "[class*='number-contact']",
        ]:
            el = card.query_selector(sel)
            if not el:
                continue
            href = el.get_attribute("href") or ""
            txt = (el.inner_text() or "").strip()
            cand = href.replace("tel:", "") if href.startswith("tel:") else txt
            tel = normalize_phone(cand)
            if tel:
                break

    # Stratégie 3 : regex sur le texte complet de la carte
    if not tel:
        m = PHONE_RE.search(text_blob)
        if m:
            tel = normalize_phone(m.group(0))

    # Adresse
    adresse = ""
    for sel in ["a.adresse", ".bi-address", "address", ".adresse-container"]:
        el = card.query_selector(sel)
        if el:
            adresse = (el.inner_text() or "").strip().replace("\n", " ")
            break

    # Email (rare)
    email = ""
    mailto = card.query_selector("a[href^='mailto:']")
    if mailto:
        href = mailto.get_attribute("href") or ""
        email = href.replace("mailto:", "").strip()
    elif text_blob:
        m = EMAIL_RE.search(text_blob)
        if m:
            email = m.group(0)

    return {
        "nom": nom,
        "type": type_label,
        "telephone": tel,
        "email": email,
        "adresse": adresse,
        "ville": ville,
        "departement": dept,
        "source": "Pages Jaunes",
    }


# ─── GOOGLE MAPS ──────────────────────────────────────────────────────────────
def scrape_google_maps(page, term: str, ville: str, dept: str, type_label: str) -> list[dict]:
    """Scrape Google Maps : scroll feed + clic chaque fiche pour tél/adresse."""
    leads: list[dict] = []
    search = f"{term} {ville}".replace(" ", "+")
    url = f"https://www.google.com/maps/search/{search}?hl=fr"

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        polite_sleep(2.5, 4.0)
    except Exception as e:
        print(f"   ⚠️  GMaps load error {ville}/{term}: {e}")
        return leads

    # Accepte cookies / consent
    for sel in [
        "button:has-text('Tout accepter')",
        "button:has-text('Accepter')",
        "button[aria-label='Tout accepter']",
        "[aria-label='Accept all']",
    ]:
        try:
            btn = page.query_selector(sel)
            if btn:
                btn.click()
                polite_sleep(0.8, 1.5)
                break
        except Exception:
            pass

    # Attend la liste de résultats
    try:
        page.wait_for_selector("[role='feed']", timeout=8000)
    except Exception:
        return leads

    # Scroll dans le panneau de gauche pour charger plus de résultats
    feed = page.query_selector("[role='feed']")
    if feed:
        for _ in range(4):
            try:
                page.evaluate("(el) => el.scrollBy(0, 1500)", feed)
            except Exception:
                pass
            polite_sleep(0.7, 1.4)

    # Récupère les cartes
    cards = page.query_selector_all("[role='feed'] > div > div[role='article'], [role='feed'] > div > a.hfpxzc")
    if not cards:
        cards = page.query_selector_all("a.hfpxzc")

    seen = set()
    for card in cards[:15]:
        try:
            lead = _gmaps_extract_one(page, card, ville, dept, type_label)
            if lead and slugify(lead["nom"]) not in seen:
                seen.add(slugify(lead["nom"]))
                leads.append(lead)
        except Exception:
            continue

    return leads


def _gmaps_extract_one(page, card, ville: str, dept: str, type_label: str) -> dict | None:
    # Nom + URL
    nom = ""
    try:
        nom_el = card.query_selector(".fontHeadlineSmall, .qBF1Pd")
        if nom_el:
            nom = (nom_el.inner_text() or "").strip()
        if not nom:
            aria = card.get_attribute("aria-label") or ""
            nom = aria.strip()
    except Exception:
        pass
    if not nom or len(nom) < 3:
        return None

    # Tente d'attraper tél / adresse depuis la liste sans cliquer
    tel = ""
    adresse = ""
    try:
        text = card.inner_text() or ""
        m = PHONE_RE.search(text)
        if m:
            tel = normalize_phone(m.group(0))
    except Exception:
        pass

    # Si pas de tél, on clique sur la fiche pour ouvrir le panneau
    if not tel:
        try:
            card.click()
            page.wait_for_timeout(1800)
            # tél depuis le bouton dédié
            tel_btn = page.query_selector("button[data-item-id^='phone']")
            if tel_btn:
                aria = tel_btn.get_attribute("aria-label") or ""
                tel_txt = aria.split(":")[-1] if ":" in aria else aria
                tel = normalize_phone(tel_txt)
            # adresse
            adr_btn = page.query_selector("button[data-item-id='address']")
            if adr_btn:
                aria = adr_btn.get_attribute("aria-label") or ""
                adresse = aria.split(":")[-1].strip() if ":" in aria else aria.strip()
        except Exception:
            pass

    if not tel and not adresse:
        return None

    return {
        "nom": nom,
        "type": type_label,
        "telephone": tel,
        "email": "",
        "adresse": adresse,
        "ville": ville,
        "departement": dept,
        "source": "Google Maps",
    }


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────
def write_excel(leads: list[dict], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    headers = ["Entreprise", "Type", "Téléphone", "Email", "Adresse", "Ville", "Département", "Source"]
    keymap = {
        "Entreprise": "nom", "Type": "type", "Téléphone": "telephone",
        "Email": "email", "Adresse": "adresse", "Ville": "ville",
        "Département": "departement", "Source": "source",
    }

    wb = Workbook()
    sheets = [
        ("Tous les leads", leads),
        ("Syndics", [l for l in leads if l["type"] == "Syndic"]),
        ("Agences", [l for l in leads if l["type"] == "Agence"]),
    ]

    navy_fill = PatternFill("solid", fgColor="0D1F38")
    red_fill = PatternFill("solid", fgColor="CC2020")
    alt_fill = PatternFill("solid", fgColor="F5F5F3")
    white_font = Font(color="FFFFFF", bold=True, name="Calibri")
    border = Side(border_style="thin", color="E0E0DB")
    cell_border = Border(left=border, right=border, top=border, bottom=border)

    first = True
    for sheet_name, rows in sheets:
        ws = wb.active if first else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        first = False

        # Bandeau Belmonts
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = ws.cell(row=1, column=1, value="BELMONTS — Leads commerciaux IDF")
        cell.fill = red_fill
        cell.font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Header
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = navy_fill
            c.font = white_font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = cell_border
        ws.row_dimensions[2].height = 22

        # Data
        for i, lead in enumerate(rows, start=3):
            for col, h in enumerate(headers, 1):
                v = lead.get(keymap[h], "")
                c = ws.cell(row=i, column=col, value=v)
                c.border = cell_border
                if i % 2 == 0:
                    c.fill = alt_fill

        # Largeurs
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = max(
                [len(str(headers[col - 1]))]
                + [len(str(r.get(keymap[headers[col - 1]], ""))) for r in rows[:300]]
            )
            ws.column_dimensions[letter].width = min(max_len + 3, 55)

        ws.freeze_panes = "A3"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ─── ORCHESTRATION ────────────────────────────────────────────────────────────
def dedupe(leads: list[dict]) -> list[dict]:
    """
    Dédup par TÉLÉPHONE en priorité — un téléphone = un bureau physique unique.
    Indispensable pour les enseignes type Foncia / Citya / Century21 qui ont
    des dizaines de bureaux différents en IDF (chacun = un prospect distinct).

    Fallback : (nom + adresse) pour les rares leads sans téléphone.
    Les leads sans tél ni adresse sont écartés (inutilisables pour prospecter).
    """
    by_key: dict[str, dict] = {}
    for l in leads:
        tel = (l.get("telephone") or "").strip()
        nom = (l.get("nom") or "").strip()
        adresse = (l.get("adresse") or "").strip()

        if tel:
            key = f"tel:{tel}"
        elif nom and adresse:
            key = f"na:{slugify(nom)}|{slugify(adresse)}"
        else:
            continue  # pas de tél, pas d'adresse → lead non exploitable

        existing = by_key.get(key)
        if not existing:
            by_key[key] = dict(l)
            continue
        # Fusion : on garde les champs non vides
        for k in ("telephone", "email", "adresse"):
            if not existing.get(k) and l.get(k):
                existing[k] = l[k]
        sources = {existing.get("source", ""), l.get("source", "")}
        existing["source"] = " + ".join(sorted(s for s in sources if s))
    return list(by_key.values())


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Scraper Belmonts — leads IDF")
    p.add_argument("--zone", default="all",
                   help="Codes dpt séparés par virgule (ex: 75,92) ou 'all' (défaut)")
    p.add_argument("--test", action="store_true", help="Mode test : 1 arrondissement Paris")
    p.add_argument("--no-google", action="store_true", help="Pages Jaunes seul")
    p.add_argument("--no-pj", action="store_true", help="Google Maps seul")
    p.add_argument("--headless", action="store_true", help="Navigateur invisible")
    p.add_argument("--output", default=None, help="Chemin du fichier Excel de sortie")
    p.add_argument("--reset", action="store_true", help="Ignore le cache de reprise")
    return p.parse_args()


def select_villes(args) -> list[tuple[str, str]]:
    """Retourne la liste (dept, ville) à scraper selon les args."""
    if args.test:
        return [("75", "Paris 1")]
    if args.zone == "all":
        depts = list(ZONES.keys())
    else:
        depts = [d.strip() for d in args.zone.split(",") if d.strip() in ZONES]
    villes: list[tuple[str, str]] = []
    for d in depts:
        for v in ZONES[d]:
            villes.append((d, v))
    return villes


def fmt_eta(seconds: float) -> str:
    if seconds < 60:
        return f"{int(seconds)}s"
    m = int(seconds // 60)
    s = int(seconds % 60)
    return f"{m}min{s:02d}"


def run() -> int:
    args = parse_args()
    if args.reset and CACHE_FILE.exists():
        CACHE_FILE.unlink()

    cache = load_cache()
    done_set = {tuple(x) for x in cache.get("done", [])}
    all_leads: list[dict] = list(cache.get("leads", []))

    villes = select_villes(args)
    if not villes:
        print("❌ Aucune ville à scraper. Vérifie --zone.")
        return 1

    iters: list[tuple[str, str, dict]] = []
    for dept, ville in villes:
        for q in QUERIES:
            iters.append((dept, ville, q))
    total = len(iters)

    out_path = Path(args.output) if args.output else (
        Path(__file__).resolve().parent / f"leads_belmonts_{datetime.now():%Y%m%d}.xlsx"
    )

    print(f"\n🏗️  BELMONTS — scraping de {len(villes)} ville(s) × {len(QUERIES)} types = {total} recherches")
    print(f"📁 Sortie : {out_path}")
    print(f"🔧 PJ={not args.no_pj}  GMaps={not args.no_google}  Headless={args.headless}")
    if done_set:
        print(f"♻️  Reprise depuis cache : {len(done_set)}/{total} déjà fait")
    print("─" * 70)

    start = time.time()
    new_count = 0

    # Si tout est déjà en cache, on saute le lancement de Chromium et on va
    # direct à la dédup + Excel. Évite 5s d'attente inutile.
    todo = [(d, v, q) for d, v, q in iters if (d, v, q["label"]) not in done_set]
    if not todo:
        print("✨ Tout est déjà scrapé (cache complet) — on passe direct à l'Excel.\n")
    else:
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            print("❌ Playwright n'est pas installé. Lance : pip3 install -r requirements.txt && playwright install chromium")
            return 1

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=args.headless)
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
                ),
                locale="fr-FR",
                viewport={"width": 1280, "height": 900},
            )
            page = context.new_page()

            for idx, (dept, ville, q) in enumerate(iters, 1):
                key = (dept, ville, q["label"])
                if key in done_set:
                    continue

                elapsed = time.time() - start
                done_now = idx - 1
                eta = (elapsed / done_now * (total - done_now)) if done_now else 0

                print(f"\n[{idx}/{total}] {ville} ({dept}) — {q['label']}   ⏱ ETA: {fmt_eta(eta)}")

                batch: list[dict] = []
                if not args.no_pj:
                    try:
                        pj = scrape_pages_jaunes(page, q["term"], ville, dept, q["label"])
                        print(f"   📒 Pages Jaunes : {len(pj)} fiches  (avec tel: {sum(1 for l in pj if l['telephone'])})")
                        batch.extend(pj)
                    except Exception as e:
                        print(f"   ⚠️  PJ erreur : {e}")
                    polite_sleep()

                if not args.no_google:
                    try:
                        gm = scrape_google_maps(page, q["term"], ville, dept, q["label"])
                        print(f"   🗺️  Google Maps  : {len(gm)} fiches  (avec tel: {sum(1 for l in gm if l['telephone'])})")
                        batch.extend(gm)
                    except Exception as e:
                        print(f"   ⚠️  GMaps erreur : {e}")
                    polite_sleep()

                all_leads.extend(batch)
                new_count += len(batch)
                done_set.add(key)

                # Sauvegarde du cache toutes les 5 itérations
                if idx % 5 == 0:
                    cache["done"] = [list(k) for k in done_set]
                    cache["leads"] = all_leads
                    save_cache(cache)

            browser.close()

    cache["done"] = [list(k) for k in done_set]
    cache["leads"] = all_leads
    save_cache(cache)

    # Dédup et export Excel
    print("\n" + "─" * 70)
    print(f"🔁 Déduplication de {len(all_leads)} fiches…")
    final_leads = dedupe(all_leads)
    print(f"✅ {len(final_leads)} leads uniques ({sum(1 for l in final_leads if l['telephone'])} avec téléphone)")

    print(f"💾 Écriture du fichier Excel…")
    write_excel(final_leads, out_path)

    duration = time.time() - start
    print(f"\n🎉 Terminé en {fmt_eta(duration)}")
    print(f"📂 Fichier : {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(run())
